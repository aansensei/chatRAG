from pathlib import Path
from openpyxl import load_workbook
from .base import ExtractResult


def extract_xlsx(file_path: str) -> ExtractResult:
    path = Path(file_path)
    result = ExtractResult()
    wb = load_workbook(file_path, data_only=True)
    tables = []
    text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # skip empty sheets
        non_empty = [r for r in rows if any(c is not None for c in r)]
        if not non_empty:
            continue

        header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(non_empty[0])]
        data = [
            {header[i]: cell for i, cell in enumerate(row)}
            for row in non_empty[1:]
        ]
        tables.append({
            "sheet": sheet_name,
            "data": data
        })

        # Flat text per sheet for embedding. Joined with a blank line between
        # rows (not a single newline) so the chunker's paragraph-boundary
        # splitting — which only breaks on "\n\n" — treats each row as an
        # atomic unit instead of packing the whole sheet into one run-on
        # paragraph. Without this, a sheet that exceeds the chunk token
        # budget gets sliced by the chunker's word-level fallback, which can
        # cut a row (or even a single cell's number) in half; a small/medium
        # table (the common case for KPI/budget sheets) now fits in one
        # chunk intact, so sum/average questions over it see every row.
        sheet_lines = []
        for row in non_empty:
            # Two spaces around the separator, not one — matches
            # office_extractor.py's docx/csv table formatting, which
            # ask_question.py's has_tabular detection ("  |  ") relies on to
            # keep a table's chunks intact instead of putting them through
            # relevance reranking that could drop/reorder rows.
            line = "  |  ".join(str(c) for c in row if c is not None)
            if line.strip():
                sheet_lines.append(line)
        if sheet_lines:
            text_parts.append("\n\n".join(sheet_lines))

    result.text = "\n\n".join(text_parts)
    result.tables = tables
    result.metadata = {"sheets": wb.sheetnames, "source": str(path)}

    return result
